#!/usr/bin/python3

from time import sleep
import requests
import openpyxl
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

options = Options()
#options.headless = True
driver = webdriver.Firefox(service_log_path='/dev/null',options=options)
login = 'olegvolov988@gmail.com'
password = 'gfhjkm123'


driver.get('https://www.strava.com/login')
#https://www.strava.com/athletes/52373249
main_page = driver.page_source























#print(main_page)

element_text = driver.find_element_by_id("email")
element_text.click()
element_text.send_keys(login)

element_p = driver.find_element_by_id("password")
element_p.click()
element_p.send_keys(password)

driver.find_element_by_id("login-button").click()

num_pages = 80
hrefs = []

for i in range(2,num_pages+1):
    
    url_str = "https://www.strava.com/clubs/269512/members?page={}".format(i)
    driver.get(url_str)
    #print(url_str)
    html_list = driver.find_elements_by_class_name("list-athletes")
    items = html_list[1].find_elements_by_tag_name("li")
    
    for item in items:
        href = str(item.find_element_by_class_name("text-headline").find_element_by_tag_name("a").get_attribute("href"))
        hrefs.append(href)
        

row = 2 #rows in excel
wb = openpyxl.Workbook()
wb.create_sheet(title = "Times", index = 0)
sheet = wb["Times"]
for h in hrefs:
    driver.get(h)
    print(h)
    try:
        table1 = driver.find_element(By.XPATH, "//*[contains(@class,'running')]").find_element(By.XPATH, "//table[contains(@class,'dense')]/tbody[3]/tr[6]/td[1]")
        table2 = driver.find_element(By.XPATH, "/html/body/div[2]/div[3]/div/div[2]/div[2]/div[3]/div[2]/table/tbody[3]/tr[6]/td[2]/a")
        tablem = driver.find_element(By.XPATH, "//*[contains(@class,'running')]").find_element(By.XPATH, "//table[contains(@class,'dense')]/tbody[3]/tr[7]/td[2]/a")
        tableh = driver.find_element(By.XPATH, "//*[contains(@class,'running')]").find_element(By.XPATH, "//table[contains(@class,'dense')]/tbody[3]/tr[6]/td[2]/a")
        
        tablel = driver.find_element(By.XPATH, "//*[contains(@class,'running')]").find_elements(By.XPATH, "//table[contains(@class,'dense')]/tbody[3]/tr[6]/td")
        

        #print(table1.get_attribute("innerHTML").splitlines()[0])
        th = tableh.get_attribute("innerHTML").splitlines()[0]
        print(th) # > to A
        tm = tablem.get_attribute("innerHTML").splitlines()[0]
        print(tm) # > to B
        cellh = sheet.cell(row = row, column = 1)
        cellh.value = th
        cellm = sheet.cell(row = row, column = 2)
        cellm.value = tm
        #=МИНУТЫ(A2)*60+СЕКУНДЫ(A2) + ЦЕЛОЕ(A2*24)*60*60
        cells = sheet.cell(row = row, column = 3)
        cells.value = h
        row += 1
        
    except:
        pass
    
    #print("done")

wb.save('time.xlsx')
'''
        table1 = driver.find_element(By.XPATH, "//*[contains(@class,'running')]").find_element(By.XPATH, "//table[contains(@class,'dense')]/tbody[3]/tr[6]/td[1]")
        table2 = driver.find_element(By.XPATH, "/html/body/div[2]/div[3]/div/div[2]/div[2]/div[3]/div[2]/table/tbody[3]/tr[6]/td[2]/a")
        tableh = driver.find_element(By.XPATH, "//*[contains(@class,'running')]").find_element(By.XPATH, "//table[contains(@class,'dense')]/tbody[3]/tr[6]/td[2]/a")
        tablem = driver.find_element(By.XPATH, "//*[contains(@class,'running')]").find_element(By.XPATH, "//table[contains(@class,'dense')]/tbody[3]/tr[7]/td[2]")
        tablel = driver.find_element(By.XPATH, "//*[contains(@class,'running')]").find_elements(By.XPATH, "//table[contains(@class,'dense')]/tbody[3]/tr[6]/td")

        print(table1.get_attribute("innerHTML").splitlines()[0])
        #print(table2.text)

        print(tableh.text)
        
        print(tablem.text)
        try:
            pass
        except:
            pass
'''


'''
        print(len(tablel))
        for t in range(1, len(tablel)+1):
            print(driver.find_element(By.XPATH, "//*[contains(@class,'running')]").find_element(By.XPATH, "//table[contains(@class,'dense')]/tbody[3]/tr[6]/td["+str(i)+"]").text)
'''

# element_text = driver.find_element_by_id("running-ytd")

